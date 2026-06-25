"""
migrate_alsari_db.py
====================
Lee Alsari_DB.xlsx (raíz del monorepo) y migra los datos maestros a Supabase.

Tablas destino:
  - sociedades
  - proyectos
  - kpis_sociedades
  - kpis_proyectos

Uso:
  cd <raíz del monorepo>
  pip install openpyxl requests
  python scripts/migrate_alsari_db.py

Requiere:
  - SUPABASE_URL  → URL del proyecto (o se usa el hardcoded de abajo)
  - SUPABASE_KEY  → anon key (o se usa el hardcoded de abajo)
  Se pueden sobreescribir con variables de entorno del mismo nombre.
"""

import os
import datetime
import openpyxl
import requests

# ── Configuración Supabase ──────────────────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://swtyxysvnfcfxziclteq.supabase.co")
SUPABASE_KEY = os.getenv(
    "SUPABASE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InN3dHl4eXN2bmZjZnh6aWNsdGVxIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY0MzA4NzMsImV4cCI6MjA5MjAwNjg3M30.aSHBdsXyT-TAdA7rv79qZFbLyMjVd25kX1E0GwBZTQ8",
)
XLSX_PATH = os.getenv("XLSX_PATH", "Alsari_DB.xlsx")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates",  # upsert
}

# ── Helpers ─────────────────────────────────────────────────────────────────

def _id(raw) -> str | None:
    """Convierte IDs de Excel a string limpio. 1.0 → '1', 'S-001' → 'S-001'."""
    if raw is None:
        return None
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def _num(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    try:
        return float(str(raw).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _date(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (datetime.datetime, datetime.date)):
        return raw.strftime("%Y-%m-%d")
    return None


def upsert(table: str, rows: list[dict]) -> None:
    if not rows:
        print(f"  {table}: sin filas, omitido.")
        return
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    res = requests.post(url, json=rows, headers=HEADERS, timeout=30)
    if res.status_code not in (200, 201):
        raise RuntimeError(f"Error upsert {table}: {res.status_code} — {res.text[:300]}")
    print(f"  {table}: {len(rows)} filas insertadas/actualizadas.")


# ── Lectura Excel ────────────────────────────────────────────────────────────

def read_wb(path: str):
    print(f"Abriendo {path} …")
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def parse_sociedades(wb) -> list[dict]:
    ws = wb["Sociedades"]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) < 2:
            continue
        raw_id = row[0]
        if raw_id is None:
            continue
        id_ref = _id(raw_id)
        nombre = str(row[1]).strip() if row[1] else None
        if not nombre:
            continue
        rows.append({
            "id_ref":            id_ref,
            "nombre":            nombre,
            "cif":               str(row[2]).strip() if row[2] else None,
            "holding_principal": str(row[3]).strip() if row[3] else None,
            "pct_pavier":        _num(row[4]),
            "pct_armia":         _num(row[5]),
            "estado":            str(row[6]).strip() if row[6] else None,
        })
    return rows


def parse_proyectos(wb) -> list[dict]:
    ws = wb["Proyectos"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        raw_id = row[0]
        if raw_id is None:
            continue
        id_ref = _id(raw_id)
        nombre = str(row[1]).strip() if row[1] else None
        if not nombre:
            continue
        sociedad = _id(row[2]) if row[2] else None
        rows.append({
            "id_ref":            id_ref,
            "nombre":            nombre,
            "sociedad_tenedora": sociedad,
            "estado":            str(row[3]).strip() if row[3] else None,
        })
    return rows


def parse_kpis_sociedades(wb) -> list[dict]:
    ws = wb["Datos_Financieros_Sociedades"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        raw_id = row[0]
        if raw_id is None:
            continue
        id_ref = _id(raw_id)
        nombre = str(row[1]).strip() if row[1] else None
        if not id_ref or not nombre:
            continue
        rows.append({
            "id_ref":                id_ref,
            "nombre":                nombre,
            "tipo":                  str(row[2]).strip() if row[2] else None,
            "caja_disponible":       _num(row[3]),
            "deuda_bancaria_cp":     _num(row[4]),
            "deuda_bancaria_lp":     _num(row[5]),
            "deuda_bancaria":        _num(row[6]),
            "deuda_socios":          _num(row[7]),
            "deuda_financiera_neta": _num(row[8]),
            "activo_corriente":      _num(row[9]),
            "activo_no_corriente":   _num(row[10]),
            "activo_total":          _num(row[11]),
            "pasivo_corriente":      _num(row[12]),
            "pasivo_no_corriente":   _num(row[13]),
            "pasivo_total":          _num(row[14]),
            "fondo_maniobra":        _num(row[15]),
            "patrimonio_neto":       _num(row[16]),
            "fecha_actualizacion":   _date(row[17]),
        })
    return rows


def parse_kpis_proyectos(wb, proyectos: list[dict]) -> list[dict]:
    # Mapa id_ref → sociedad_tenedora para enriquecer kpis
    soc_map = {p["id_ref"]: p["sociedad_tenedora"] for p in proyectos}

    ws = wb["Datos_Financieros_Proyectos"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        raw_id = row[0]
        if raw_id is None:
            continue
        id_ref = _id(raw_id)
        nombre = str(row[1]).strip() if row[1] else None
        if not id_ref or not nombre:
            continue
        rows.append({
            "id_ref":                    id_ref,
            "nombre":                    nombre,
            "tipo":                      str(row[2]).strip() if row[2] else None,
            "sociedad_tenedora":         soc_map.get(id_ref),
            "presupuesto":               _num(row[3]),
            "inversion_inicial":         _num(row[4]),
            "recapex_acumulado":         _num(row[5]),
            "capital_expuesto":          _num(row[6]),
            "deuda":                     _num(row[7]),
            "beneficio_acumulado_bruto": _num(row[8]),
            "opex_acumulado":            _num(row[9]),
            "valoracion":                _num(row[10]),
            "margen_latente_salida":     _num(row[11]),
            "margen_latente_salida_pct": _num(row[12]),
            "fecha_actualizacion":       _date(row[13]),
        })
    return rows


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = read_wb(XLSX_PATH)

    print("\n[1/4] Parseando Sociedades …")
    soc = parse_sociedades(wb)
    print(f"      {len(soc)} sociedades encontradas.")

    print("[2/4] Parseando Proyectos …")
    proj = parse_proyectos(wb)
    print(f"      {len(proj)} proyectos encontrados.")

    print("[3/4] Parseando KPIs Sociedades …")
    kpis_soc = parse_kpis_sociedades(wb)
    print(f"      {len(kpis_soc)} filas de KPIs encontradas.")

    print("[4/4] Parseando KPIs Proyectos …")
    kpis_proj = parse_kpis_proyectos(wb, proj)
    print(f"      {len(kpis_proj)} filas de KPIs proyectos encontradas.")

    wb.close()

    print("\nSubiendo a Supabase …")
    upsert("sociedades",      soc)
    upsert("proyectos",       proj)
    upsert("kpis_sociedades", kpis_soc)
    upsert("kpis_proyectos",  kpis_proj)

    print("\nMigracion completada OK.")


if __name__ == "__main__":
    main()
