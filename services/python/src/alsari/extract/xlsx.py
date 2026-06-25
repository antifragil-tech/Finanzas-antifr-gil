"""Extracción de Excel (.xlsx) y Google Sheets exportados.

Convierte cada hoja a una tabla Markdown legible para el LLM y el chunker.
"""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook


def extract_text(data: bytes) -> str:
    """Lee un .xlsx y devuelve cada hoja como tabla Markdown con cabecera."""
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    sheets: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [
            [("" if cell is None else str(cell)) for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        rows = [r for r in rows if any(c.strip() for c in r)]
        if not rows:
            continue
        sheets.append(f"## Hoja: {sheet_name}\n\n{_rows_to_markdown(rows)}")
    wb.close()
    return "\n\n".join(sheets)


def extract_from_csv(data: bytes) -> str:
    """Google Sheets exporta CSV (una hoja). Lo convertimos también a Markdown."""
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return ""
    return _rows_to_markdown(rows)


def _rows_to_markdown(rows: list[list[str]]) -> str:
    width = max(len(r) for r in rows)
    padded = [r + [""] * (width - len(r)) for r in rows]
    header, *body = padded
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)
